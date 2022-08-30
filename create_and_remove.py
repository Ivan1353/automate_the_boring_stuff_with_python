import openpyxl
wb = openpyxl.Workbook()
wb.sheetnames
wb.create_sheet()
wb.sheetnames
wb.create_sheet(index=0, title='First Sheet')
wb.create_sheet(index=2, title='Middle Sheet'
del wb['Middle Sheet']
del wb['Sheet1']
wb.save('xlfile.xlsx')