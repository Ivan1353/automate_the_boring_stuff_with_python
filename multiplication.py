import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb['Sheet']
bold = Font(bold=True)
num = int(input())

for n in range(2, num+2):

    the_row = sheet.cell(row = 1, column = n)
    the_row.value = n-1
    the_row.font = bold
    the_col = sheet.cell(row = n, column = 1)
    the_col.value = n-1
    the_col.font = bold

for row in range(1, num+1):
    for col in range(1, num+1):
        cell = sheet.cell(row=row+1, column=col+1)
        cell.value = row*col

wb.save('multiply.xlsx')
